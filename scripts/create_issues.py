#!/usr/bin/env python3
"""批量创建 GitHub Issues 脚本"""

import subprocess
import json
from openpyxl import load_workbook

def run_gh_command(args):
    """执行 gh 命令"""
    result = subprocess.run(['gh'] + args, capture_output=True, text=True)
    return result.returncode == 0, result.stdout, result.stderr

def create_issue(title, body, labels):
    """创建单个 Issue"""
    cmd = ['issue', 'create', '--title', title, '--body', body]
    for label in labels:
        cmd.extend(['--label', label])
    success, stdout, stderr = run_gh_command(cmd)
    if success:
        # 提取 issue URL
        url = stdout.strip()
        print(f"✅ Created: {title[:50]}... -> {url}")
        return url
    else:
        print(f"❌ Failed: {title[:50]}... -> {stderr}")
        return None

def load_backlog(xlsx_path):
    """加载 Backlog 数据"""
    wb = load_workbook(xlsx_path)

    # 加载 Epics
    epics = []
    ws = wb['Epics']
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # Epic ID 存在
            epic = dict(zip(headers, row))
            epics.append(epic)

    # 加载 Stories
    stories = []
    ws = wb['Stories']
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # Story ID 存在
            story = dict(zip(headers, row))
            stories.append(story)

    return epics, stories

def get_labels_for_epic(epic):
    """根据 Epic 数据生成 labels"""
    labels = ['type:epic']

    # Priority
    priority = epic.get('Priority', 'P0')
    labels.append(f'priority:{priority}')

    # Release
    release = epic.get('Target Release', 'v0')
    labels.append(f'release:{release}')

    # Component
    components = epic.get('Components', '')
    if components:
        comp_map = {
            'Infra': 'infra', 'Data': 'data', 'Backend': 'backend',
            'Web': 'web', 'UE': 'ue', 'QA': 'qa', 'All': 'infra'
        }
        for key, val in comp_map.items():
            if key in components:
                labels.append(f'component:{val}')
                break

    return labels

def get_labels_for_story(story):
    """根据 Story 数据生成 labels"""
    labels = []

    # Type
    story_type = story.get('Type', 'Feature').lower()
    type_map = {'feature': 'feature', 'task': 'task', 'spike': 'spike', 'test': 'test', 'doc': 'doc'}
    labels.append(f"type:{type_map.get(story_type, 'feature')}")

    # Priority
    priority = story.get('Priority', 'P0')
    labels.append(f'priority:{priority}')

    # Release
    release = story.get('Target Release', 'v0')
    labels.append(f'release:{release}')

    # Component
    component = story.get('Component', '')
    if component:
        comp_map = {
            'Infra': 'infra', 'Data': 'data', 'Backend': 'backend',
            'Web': 'web', 'UE': 'ue', 'QA': 'qa'
        }
        for key, val in comp_map.items():
            if key in component:
                labels.append(f'component:{val}')
                break

    return labels

def build_epic_body(epic, stories):
    """构建 Epic Issue 的 body"""
    epic_id = epic.get('Epic ID', '')
    epic_stories = [s for s in stories if s.get('Epic ID') == epic_id]

    body = f"""## Description
{epic.get('Description', '')}

## Components
{epic.get('Components', '')}

## Dependencies
{epic.get('Dependencies', 'None')}

## Target Release
{epic.get('Target Release', 'v0')}

## Owner
{epic.get('Owner', 'TBD')}

## Notes
{epic.get('Notes', '')}

---

## Stories ({len(epic_stories)})

"""
    for s in epic_stories:
        body += f"- [ ] **{s.get('Story ID')}**: {s.get('Title', '')}\n"

    return body

def build_story_body(story):
    """构建 Story Issue 的 body"""
    body = f"""## User Story
{story.get('User Story', '')}

## Description / Tasks
{story.get('Description / Tasks', '')}

## Acceptance Criteria
{story.get('Acceptance Criteria', '')}

## Estimate
{story.get('Estimate (SP)', '?')} Story Points

## Epic
{story.get('Epic ID', '')}

## Dependencies
{story.get('Dependencies', 'None')}
"""
    return body

def main():
    xlsx_path = 'DigitalEarth_PRD_SPEC_Backlog_v1.0/数字地球气象可视化平台_Backlog_v1.0.xlsx'

    print("📂 Loading Backlog...")
    epics, stories = load_backlog(xlsx_path)
    print(f"   Found {len(epics)} Epics, {len(stories)} Stories")

    # 创建 Epics
    print("\n📌 Creating Epic Issues...")
    epic_urls = {}
    for epic in epics:
        epic_id = epic.get('Epic ID', '')
        title = f"[{epic_id}] {epic.get('Epic Name', '')}"
        body = build_epic_body(epic, stories)
        labels = get_labels_for_epic(epic)

        url = create_issue(title, body, labels)
        if url:
            epic_urls[epic_id] = url

    print(f"\n✅ Created {len(epic_urls)} Epic Issues")

    # 创建 Stories
    print("\n📝 Creating Story Issues...")
    story_count = 0
    for story in stories:
        story_id = story.get('Story ID', '')
        title = f"[{story_id}] {story.get('Title', '')}"
        body = build_story_body(story)
        labels = get_labels_for_story(story)

        url = create_issue(title, body, labels)
        if url:
            story_count += 1

    print(f"\n✅ Created {story_count} Story Issues")
    print(f"\n🎉 Total: {len(epic_urls)} Epics + {story_count} Stories")

if __name__ == '__main__':
    main()
